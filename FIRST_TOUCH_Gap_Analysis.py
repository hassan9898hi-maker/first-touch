from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ═══════ COLORS ═══════
NAVY = "0D1B2A"
OCEAN = "1565C0"
WHITE = "FFFFFF"
LIGHT_BLUE = "D6E4F0"
LIGHT_GREEN = "C6EFCE"
LIGHT_RED = "FFC7CE"
LIGHT_YELLOW = "FFFFCC"
LIGHT_PURPLE = "E8D5F5"
LIGHT_GRAY = "F2F2F2"
GREEN = "006100"
RED = "9C0006"
AMBER = "9C6500"
PURPLE = "7030A0"

hFont = Font(name="Arial", bold=True, size=14, color=WHITE)
h2Font = Font(name="Arial", bold=True, size=12, color=NAVY)
h3Font = Font(name="Arial", bold=True, size=11, color=OCEAN)
bFont = Font(name="Arial", size=10)
bBold = Font(name="Arial", size=10, bold=True)
navyFill = PatternFill("solid", fgColor=NAVY)
oceanFill = PatternFill("solid", fgColor=OCEAN)
greenFill = PatternFill("solid", fgColor=LIGHT_GREEN)
redFill = PatternFill("solid", fgColor=LIGHT_RED)
yellowFill = PatternFill("solid", fgColor=LIGHT_YELLOW)
purpleFill = PatternFill("solid", fgColor=LIGHT_PURPLE)
grayFill = PatternFill("solid", fgColor=LIGHT_GRAY)
blueFill = PatternFill("solid", fgColor=LIGHT_BLUE)
centerA = Alignment(horizontal="center", vertical="center", wrap_text=True)
rightA = Alignment(horizontal="right", vertical="center")
leftA = Alignment(horizontal="left", vertical="center", wrap_text=True)
thin = Side(style="thin", color="999999")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def style_header_row(ws, row, cols, fill=oceanFill):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(name="Arial", bold=True, size=10, color=WHITE)
        cell.fill = fill
        cell.alignment = centerA
        cell.border = border

def style_data_row(ws, row, cols, fill=None):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = bFont
        cell.alignment = leftA if c <= 2 else centerA
        cell.border = border
        if fill: cell.fill = fill

def priority_fill(p):
    if p == "حرج": return redFill
    if p == "عالي": return PatternFill("solid", fgColor="FFD7D7")
    if p == "متوسط": return yellowFill
    return greenFill

# ══════════════════════════════════════════════════════════
# SHEET 1: Executive Summary
# ══════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "الملخص التنفيذي"
ws1.sheet_view.rightToLeft = True

ws1.merge_cells("A1:H1")
ws1.cell(1,1, "FIRST TOUCH — تحليل الفجوات وخطة الإغلاق").font = Font(name="Arial", bold=True, size=18, color=WHITE)
ws1.cell(1,1).fill = navyFill
ws1.cell(1,1).alignment = centerA
ws1.row_dimensions[1].height = 45

ws1.merge_cells("A2:H2")
ws1.cell(2,1, "منصة إدارة مشاريع البناء — التاريخ: 31 مارس 2026").font = Font(name="Arial", size=11, color=WHITE)
ws1.cell(2,1).fill = oceanFill
ws1.cell(2,1).alignment = centerA

# Summary stats
stats = [
    ("نسبة الإكتمال الكلية", "85%"),
    ("عدد الميزات المكتملة", "14"),
    ("عدد الفجوات الحرجة", "5"),
    ("عدد الفجوات المتوسطة", "5"),
    ("عدد الفجوات المنخفضة", "4"),
    ("إجمالي سطور الكود (Backend)", "3,413"),
    ("إجمالي سطور الكود (Frontend)", "5,321"),
    ("تغطية الاختبارات", "0%"),
]
r = 4
ws1.cell(r, 1, "المؤشر").font = bBold; ws1.cell(r, 1).fill = blueFill; ws1.cell(r,1).border = border
ws1.cell(r, 2, "القيمة").font = bBold; ws1.cell(r, 2).fill = blueFill; ws1.cell(r,2).border = border
for label, val in stats:
    r += 1
    ws1.cell(r, 1, label).font = bFont; ws1.cell(r,1).border = border
    ws1.cell(r, 2, val).font = bBold; ws1.cell(r,2).border = border; ws1.cell(r,2).alignment = centerA

# Feature matrix
r += 2
ws1.merge_cells(f"A{r}:H{r}")
ws1.cell(r, 1, "مصفوفة اكتمال الميزات").font = h2Font
ws1.cell(r, 1).fill = grayFill

r += 1
headers = ["الميزة", "Backend %", "Frontend %", "الحالة", "الأولوية", "Sprint", "الملاحظات", "المسؤول"]
for i, h in enumerate(headers, 1):
    ws1.cell(r, i, h)
style_header_row(ws1, r, 8)

features = [
    ("التسجيل والمصادقة", "100%", "100%", "مكتمل", "-", "-", "JWT + OTP + تحقق بريد", "Backend"),
    ("المشاريع (المالك)", "100%", "95%", "مكتمل", "-", "-", "إنشاء، عروض، موافقات", "Full Stack"),
    ("تحليل AI الذكي", "100%", "100%", "مكتمل", "-", "-", "Claude API + تحليل ملفات", "Backend"),
    ("عروض الأسعار (BOQ)", "100%", "90%", "مكتمل", "-", "-", "تقديم، قبول، رفض", "Full Stack"),
    ("الفحص والاعتماد", "100%", "90%", "مكتمل", "-", "-", "فحص، موافقة، رفض", "Full Stack"),
    ("تسليم أعمال المقاول", "100%", "85%", "مكتمل", "-", "-", "تسليم بنود مع توثيق", "Full Stack"),
    ("المحفظة والدفع", "100%", "80%", "مكتمل", "-", "-", "إيداع، سحب، معاملات", "Full Stack"),
    ("لوحة الأدمن", "85%", "70%", "مكتمل", "منخفض", "Sprint 3", "ينقص: طلبات البنك", "Backend"),
    ("المجمعات (المطور)", "70%", "90%", "جزئي", "متوسط", "Sprint 2", "صلاحيات، validation", "Full Stack"),
    ("رفع الملفات", "100%", "50%", "جزئي", "عالي", "Sprint 1", "Frontend غير مربوط", "Frontend"),
    ("التعليقات", "100%", "40%", "جزئي", "متوسط", "Sprint 2", "UI ناقص", "Frontend"),
    ("الإشعارات الفورية", "50%", "0%", "غير مكتمل", "حرج", "Sprint 1", "Socket.io غير متصل", "Full Stack"),
    ("التمويل", "30%", "20%", "ناقص", "متوسط", "Sprint 2", "بدون workflow", "Full Stack"),
    ("اختبارات", "0%", "0%", "معدوم", "عالي", "Sprint 3", "لا تغطية", "QA"),
]

for feat in features:
    r += 1
    for i, val in enumerate(feat, 1):
        ws1.cell(r, i, val)
    style_data_row(ws1, r, 8)
    status = feat[3]
    if status == "مكتمل": ws1.cell(r, 4).fill = greenFill
    elif status == "جزئي": ws1.cell(r, 4).fill = yellowFill
    elif status in ("غير مكتمل", "ناقص", "معدوم"): ws1.cell(r, 4).fill = redFill

ws1.column_dimensions['A'].width = 28
ws1.column_dimensions['B'].width = 14
ws1.column_dimensions['C'].width = 14
ws1.column_dimensions['D'].width = 14
ws1.column_dimensions['E'].width = 12
ws1.column_dimensions['F'].width = 12
ws1.column_dimensions['G'].width = 32
ws1.column_dimensions['H'].width = 14

# ══════════════════════════════════════════════════════════
# SHEET 2: Gap Closure Plan (Prioritized)
# ══════════════════════════════════════════════════════════
ws2 = wb.create_sheet("خطة إغلاق الفجوات")
ws2.sheet_view.rightToLeft = True

ws2.merge_cells("A1:J1")
ws2.cell(1,1, "خطة إغلاق الفجوات — مرتبة حسب الأهمية").font = Font(name="Arial", bold=True, size=16, color=WHITE)
ws2.cell(1,1).fill = navyFill; ws2.cell(1,1).alignment = centerA
ws2.row_dimensions[1].height = 40

r = 3
headers2 = ["#", "الفجوة", "الوصف التفصيلي", "الأولوية", "التأثير", "Sprint", "المدة", "تاريخ البدء", "تاريخ الانتهاء", "الحالة"]
for i, h in enumerate(headers2, 1):
    ws2.cell(r, i, h)
style_header_row(ws2, r, 10)

gaps = [
    # CRITICAL - Sprint 1 (Week 1-2)
    (1, "أمان البيانات المكشوفة", "إزالة DATABASE_URL و JWT secrets المكتوبة مباشرة في server.js و start.js واستخدام .env فقط", "حرج", "أمني — اختراق كامل", "Sprint 1", "1 يوم", "1 أبريل 2026", "1 أبريل 2026", "معلّق"),
    (2, "ربط Socket.io بالفرونت", "إنشاء اتصال Socket.io في App.jsx، الانضمام لغرفة المستخدم، استقبال الإشعارات الفورية وتحديث الـ badges", "حرج", "تجربة المستخدم — لا إشعارات فورية", "Sprint 1", "3 أيام", "1 أبريل 2026", "3 أبريل 2026", "معلّق"),
    (3, "تسجيل أرباح المقاول تلقائياً", "عند موافقة المالك على بند، إنشاء سجل ContractorEarning تلقائياً وتحديث المحفظة", "حرج", "مالي — المقاول لا يحصل على أرباحه", "Sprint 1", "2 يوم", "3 أبريل 2026", "4 أبريل 2026", "معلّق"),
    (4, "ربط رفع الملفات بالواجهة", "تحويل FileUploader لإرسال FormData بدل Base64، ربطه مع Cloudinary/Local storage، معاينة قبل الرفع", "حرج", "وظيفي — الملفات لا تُحفظ", "Sprint 1", "3 أيام", "4 أبريل 2026", "7 أبريل 2026", "معلّق"),
    (5, "إضافة endpoint طلبات البنك للأدمن", "GET /api/admin/bank-requests مع فلترة وموافقة/رفض + واجهة أدمن لإدارتها", "حرج", "وظيفي — طلبات بدون مراجعة", "Sprint 1", "2 يوم", "7 أبريل 2026", "8 أبريل 2026", "معلّق"),

    # HIGH - Sprint 1 (Week 2)
    (6, "تكامل AI Upload في الواجهة", "ربط نموذج الرفع الذكي الجديد بالـ modal، إضافة زر في dashboard المالك، اختبار التحليل", "عالي", "ميزة رئيسية — AI غير متاح للمستخدم", "Sprint 1", "2 يوم", "9 أبريل 2026", "10 أبريل 2026", "معلّق"),
    (7, "إصلاح workflow التمويل", "إضافة endpoints موافقة/رفض، جدول سداد، رفع مستندات، واجهة تتبع الحالة", "عالي", "وظيفي — ميزة معطلة", "Sprint 2", "4 أيام", "13 أبريل 2026", "16 أبريل 2026", "معلّق"),
    (8, "Pagination لجميع القوائم", "إضافة limit/offset لـ projects, compounds, notifications، عرض أرقام الصفحات", "عالي", "أداء — بطء مع بيانات كثيرة", "Sprint 2", "2 يوم", "16 أبريل 2026", "17 أبريل 2026", "معلّق"),

    # MEDIUM - Sprint 2 (Week 3-4)
    (9, "اكتمال ميزة المجمعات", "تفعيل صلاحيات المشرفين، validation على PATCH، ربط تقدم الوحدات بإحصائيات المجمع", "متوسط", "وظيفي — edge cases", "Sprint 2", "4 أيام", "17 أبريل 2026", "22 أبريل 2026", "معلّق"),
    (10, "واجهة التعليقات", "إضافة مكون عرض/إضافة التعليقات في تفاصيل البنود، تحديث فوري عبر Socket.io", "متوسط", "تجربة المستخدم", "Sprint 2", "2 يوم", "22 أبريل 2026", "23 أبريل 2026", "معلّق"),
    (11, "رفع صور مشاكل الجودة", "ربط رفع الصور في QualityIssue مع Cloudinary، معاينة الصور في الواجهة", "متوسط", "وظيفي — توثيق ناقص", "Sprint 2", "2 يوم", "23 أبريل 2026", "24 أبريل 2026", "معلّق"),
    (12, "البحث والفلترة", "إضافة بحث في المشاريع، المجمعات، المستخدمين مع فلاتر حسب الحالة والنوع", "متوسط", "تجربة المستخدم", "Sprint 2", "3 أيام", "24 أبريل 2026", "28 أبريل 2026", "معلّق"),
    (13, "تنسيق التواريخ", "إنشاء utility لتنسيق التواريخ بالعربي واستخدامه في كل الواجهات", "متوسط", "تجربة المستخدم", "Sprint 2", "1 يوم", "28 أبريل 2026", "28 أبريل 2026", "معلّق"),

    # LOW - Sprint 3 (Week 5-6)
    (14, "تصميم متجاوب (Responsive)", "إضافة media queries و breakpoints للموبايل والتابلت", "منخفض", "وصول — لا يعمل على الموبايل", "Sprint 3", "5 أيام", "29 أبريل 2026", "5 مايو 2026", "معلّق"),
    (15, "Accessibility (WCAG)", "إضافة ARIA labels، تنقل لوحة المفاتيح، تباين ألوان كافي", "منخفض", "وصول — غير متاح لذوي الإعاقة", "Sprint 3", "3 أيام", "5 مايو 2026", "7 مايو 2026", "معلّق"),
    (16, "منع الطلبات المكررة", "تعطيل الأزرار أثناء الإرسال، إضافة debounce، optimistic UI", "منخفض", "جودة — بيانات مكررة", "Sprint 3", "2 يوم", "7 مايو 2026", "8 مايو 2026", "معلّق"),
    (17, "اختبارات وحدة وتكامل", "إعداد Jest + Supertest للـ Backend، React Testing Library للفرونت", "منخفض", "جودة — لا ضمان", "Sprint 3", "5 أيام", "8 مايو 2026", "14 مايو 2026", "معلّق"),
    (18, "تقسيم App.jsx", "تقسيم الملف الواحد (5321 سطر) إلى مكونات منفصلة حسب الدور والميزة", "منخفض", "صيانة — صعوبة التطوير", "Sprint 3", "3 أيام", "14 مايو 2026", "18 مايو 2026", "معلّق"),
]

for gap in gaps:
    r += 1
    for i, val in enumerate(gap, 1):
        ws2.cell(r, i, val)
    style_data_row(ws2, r, 10)
    ws2.cell(r, 4).fill = priority_fill(gap[3])
    ws2.cell(r, 4).alignment = centerA

ws2.column_dimensions['A'].width = 5
ws2.column_dimensions['B'].width = 28
ws2.column_dimensions['C'].width = 55
ws2.column_dimensions['D'].width = 12
ws2.column_dimensions['E'].width = 30
ws2.column_dimensions['F'].width = 12
ws2.column_dimensions['G'].width = 10
ws2.column_dimensions['H'].width = 16
ws2.column_dimensions['I'].width = 16
ws2.column_dimensions['J'].width = 10

# ══════════════════════════════════════════════════════════
# SHEET 3: Sprint Timeline
# ══════════════════════════════════════════════════════════
ws3 = wb.create_sheet("الجدول الزمني")
ws3.sheet_view.rightToLeft = True

ws3.merge_cells("A1:H1")
ws3.cell(1,1, "الجدول الزمني — Sprints").font = Font(name="Arial", bold=True, size=16, color=WHITE)
ws3.cell(1,1).fill = navyFill; ws3.cell(1,1).alignment = centerA
ws3.row_dimensions[1].height = 40

sprints = [
    ("Sprint 1 — الأساسيات الحرجة", "1 - 12 أبريل 2026", "أسبوعان", [
        "إزالة البيانات المكشوفة من الكود",
        "ربط Socket.io بالفرونت اند (إشعارات فورية)",
        "تسجيل أرباح المقاول تلقائياً عند موافقة المالك",
        "ربط رفع الملفات مع Cloudinary/Local",
        "endpoint طلبات البنك + واجهة أدمن",
        "تكامل AI Upload في واجهة المالك",
    ]),
    ("Sprint 2 — التحسينات الوظيفية", "13 - 28 أبريل 2026", "أسبوعان", [
        "إصلاح workflow التمويل كاملاً",
        "Pagination لجميع القوائم",
        "اكتمال ميزة المجمعات (صلاحيات + validation)",
        "واجهة التعليقات على البنود",
        "رفع صور مشاكل الجودة",
        "البحث والفلترة في جميع القوائم",
        "تنسيق التواريخ بالعربي",
    ]),
    ("Sprint 3 — الجودة والصقل", "29 أبريل - 18 مايو 2026", "3 أسابيع", [
        "تصميم متجاوب (Responsive Design)",
        "Accessibility (WCAG 2.1 AA)",
        "منع الطلبات المكررة (Debounce)",
        "اختبارات وحدة وتكامل (Jest + Testing Library)",
        "تقسيم App.jsx إلى مكونات منفصلة",
    ]),
]

r = 3
for sprint_name, dates, duration, tasks in sprints:
    ws3.merge_cells(f"A{r}:H{r}")
    ws3.cell(r, 1, sprint_name).font = Font(name="Arial", bold=True, size=13, color=WHITE)
    ws3.cell(r, 1).fill = oceanFill; ws3.cell(r,1).alignment = centerA
    ws3.row_dimensions[r].height = 32
    r += 1
    ws3.cell(r, 1, "الفترة:").font = bBold
    ws3.cell(r, 2, dates).font = bFont
    ws3.cell(r, 4, "المدة:").font = bBold
    ws3.cell(r, 5, duration).font = bFont
    r += 1
    for i, task in enumerate(tasks, 1):
        ws3.cell(r, 1, i).font = bFont; ws3.cell(r,1).alignment = centerA; ws3.cell(r,1).border = border
        ws3.merge_cells(f"B{r}:G{r}")
        ws3.cell(r, 2, task).font = bFont; ws3.cell(r,2).border = border
        ws3.cell(r, 8, "معلّق").font = bFont; ws3.cell(r,8).alignment = centerA; ws3.cell(r,8).border = border
        ws3.cell(r, 8).fill = yellowFill
        r += 1
    r += 1

ws3.column_dimensions['A'].width = 8
ws3.column_dimensions['B'].width = 22
ws3.column_dimensions['C'].width = 20
ws3.column_dimensions['D'].width = 12
ws3.column_dimensions['E'].width = 16
ws3.column_dimensions['F'].width = 14
ws3.column_dimensions['G'].width = 14
ws3.column_dimensions['H'].width = 12

# ══════════════════════════════════════════════════════════
# SHEET 4: Product Workflow
# ══════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Workflow المنتج")
ws4.sheet_view.rightToLeft = True

ws4.merge_cells("A1:J1")
ws4.cell(1,1, "Workflow المنتج النهائي — رحلة المستخدم الكاملة").font = Font(name="Arial", bold=True, size=16, color=WHITE)
ws4.cell(1,1).fill = navyFill; ws4.cell(1,1).alignment = centerA
ws4.row_dimensions[1].height = 40

workflows = [
    ("رحلة صاحب المشروع (Owner)", "owner", [
        ("1", "التسجيل", "إنشاء حساب → تفعيل البريد → تسجيل الدخول", "مكتمل", "auth.js", "App.jsx (Login)"),
        ("2", "إنشاء مشروع (يدوي)", "اختيار نوع → مرحلة بناء → تفاصيل → شروط → ملفات → إرسال", "مكتمل", "projects.js POST /", "NewProjectForm"),
        ("3", "إنشاء مشروع (AI ذكي)", "رفع ملف + وصف → تحليل AI → مراجعة → إنشاء تلقائي", "مكتمل", "projects.js POST /ai-analyze", "AIProjectUpload"),
        ("4", "استقبال عروض الأسعار", "المقاولون يقدمون BOQ → المالك يراجع → يقبل/يرفض", "مكتمل", "projects.js POST /:id/quotation", "QuotationView"),
        ("5", "اختيار مفتش", "المفتشون يتقدمون → المالك يختار → يدفع أتعاب الفحص", "مكتمل", "projects.js POST /:id/apply-inspector", "InspectorApplyForm"),
        ("6", "متابعة التنفيذ", "المقاول يسلم بنود → المفتش يفحص → المالك يوافق ويدفع", "مكتمل", "projects.js POST /items/:id/*", "OwnerDecisionForm"),
        ("7", "إدارة المحفظة", "إيداع → دفع تلقائي عند الموافقة → سجل المعاملات", "مكتمل", "wallet.js", "WalletView"),
        ("8", "طلب تمويل بنكي", "تقديم طلب → انتظار الموافقة", "جزئي", "projects.js POST /bank-request", "ناقص"),
        ("9", "تقييم المقاول/المفتش", "بعد اكتمال المشروع → تقييم 1-5 نجوم", "مكتمل", "achievements.js", "RatingForm"),
    ]),
    ("رحلة المقاول (Contractor)", "contractor", [
        ("1", "التسجيل", "إنشاء حساب مقاول → بيانات الشركة → CR", "مكتمل", "auth.js", "App.jsx (Register)"),
        ("2", "تصفح المشاريع", "عرض المشاريع المتاحة للتسعير", "مكتمل", "projects.js GET /", "ProjectList"),
        ("3", "تقديم عرض سعر", "ملء BOQ تفصيلي → مدة → ضمان → إرسال", "مكتمل", "projects.js POST /:id/quotation", "BOQQuotationForm"),
        ("4", "استلام العقد", "عند قبول العرض → عرض تفاصيل العقد", "مكتمل", "dashboard.js GET /contract", "ContractView"),
        ("5", "تسليم الأعمال", "لكل بند: تنفيذ → توثيق بالصور → تسليم", "مكتمل", "projects.js POST /items/:id/submit", "ContractorSubmitForm"),
        ("6", "استلام الدفعات", "بعد موافقة المالك → الدفعة تُضاف للمحفظة", "جزئي — لا يُسجل تلقائياً", "wallet.js", "EarningsView"),
        ("7", "عرض الإنجازات", "المشاريع المكتملة + التقييمات", "مكتمل", "achievements.js", "AchievementsGallery"),
    ]),
    ("رحلة المفتش (Inspector)", "inspector", [
        ("1", "التسجيل", "إنشاء حساب مفتش → التخصص", "مكتمل", "auth.js", "App.jsx (Register)"),
        ("2", "التقدم للمشاريع", "تصفح → تقديم ترشيح مع الأتعاب", "مكتمل", "projects.js POST /:id/apply-inspector", "InspectorApplyForm"),
        ("3", "فحص البنود", "مراجعة تسليم المقاول → موافقة/رفض مع ملاحظات", "مكتمل", "projects.js POST /items/:id/review", "InspectorReviewForm"),
        ("4", "التقارير", "إحصائيات: نسبة الموافقة، عدد الفحوصات", "مكتمل", "dashboard.js", "InspectorDashboard"),
    ]),
    ("رحلة المطور العقاري (Developer)", "developer", [
        ("1", "إنشاء مجمع", "wizard من 6 خطوات: بيانات → وحدات → مراحل → مشرفين", "مكتمل", "compounds.js POST /", "CompoundWizard"),
        ("2", "إدارة الوحدات", "تحديث حالة كل وحدة → تتبع التقدم", "مكتمل", "compounds.js PATCH /units/:id", "UnitGrid"),
        ("3", "تتبع المواد", "إضافة مواد → تحديث الكميات → تنبيهات النقص", "مكتمل", "compounds.js /materials/*", "MaterialTracking"),
        ("4", "مشاكل الجودة", "تسجيل مشاكل → تعيين → متابعة الحل", "جزئي — بدون صور", "compounds.js /quality/*", "QualityIssues"),
        ("5", "المدفوعات", "تتبع الدفعات لكل مقاول/مرحلة", "مكتمل", "compounds.js /payments/*", "PaymentTracking"),
        ("6", "التقارير", "تقدم، مالي، gap analysis", "مكتمل", "compounds.js /reports/*", "ReportsView"),
    ]),
    ("رحلة المشرف (Admin)", "admin", [
        ("1", "إحصائيات النظام", "عدد المستخدمين، المشاريع، المحافظ", "مكتمل", "admin.js GET /stats", "AdminDashboard"),
        ("2", "إدارة المستخدمين", "بحث، تعطيل، ترقية، إعادة كلمة مرور", "مكتمل", "admin.js GET/PATCH /users/*", "UserManagement"),
        ("3", "إدارة المشاريع", "عرض جميع المشاريع مع التفاصيل", "مكتمل", "admin.js GET /projects", "ProjectManagement"),
        ("4", "طلبات البنك", "مراجعة وموافقة/رفض طلبات التمويل", "غير مكتمل", "مفقود", "مفقود"),
    ]),
]

r = 3
for wf_name, role, steps in workflows:
    ws4.merge_cells(f"A{r}:J{r}")
    colors = {"owner": "1565C0", "contractor": "E8720C", "inspector": "0EAD69", "developer": "7C3AED", "admin": "0D1B2A"}
    ws4.cell(r, 1, wf_name).font = Font(name="Arial", bold=True, size=13, color=WHITE)
    ws4.cell(r, 1).fill = PatternFill("solid", fgColor=colors.get(role, NAVY))
    ws4.cell(r, 1).alignment = centerA
    ws4.row_dimensions[r].height = 32
    r += 1

    wf_headers = ["#", "الخطوة", "الوصف", "الحالة", "Backend", "Frontend", "", "", "", ""]
    for i, h in enumerate(wf_headers, 1):
        ws4.cell(r, i, h)
    style_header_row(ws4, r, 6)
    r += 1

    for step in steps:
        for i, val in enumerate(step, 1):
            ws4.cell(r, i, val)
        style_data_row(ws4, r, 6)
        status = step[3]
        if "مكتمل" in status: ws4.cell(r, 4).fill = greenFill
        elif "جزئي" in status: ws4.cell(r, 4).fill = yellowFill
        else: ws4.cell(r, 4).fill = redFill
        r += 1
    r += 1

ws4.column_dimensions['A'].width = 5
ws4.column_dimensions['B'].width = 20
ws4.column_dimensions['C'].width = 50
ws4.column_dimensions['D'].width = 22
ws4.column_dimensions['E'].width = 30
ws4.column_dimensions['F'].width = 22

# ══════════════════════════════════════════════════════════
# SHEET 5: Risk Matrix
# ══════════════════════════════════════════════════════════
ws5 = wb.create_sheet("مصفوفة المخاطر")
ws5.sheet_view.rightToLeft = True

ws5.merge_cells("A1:G1")
ws5.cell(1,1, "مصفوفة المخاطر والتخفيف").font = Font(name="Arial", bold=True, size=16, color=WHITE)
ws5.cell(1,1).fill = navyFill; ws5.cell(1,1).alignment = centerA
ws5.row_dimensions[1].height = 40

r = 3
risk_headers = ["#", "المخاطرة", "الاحتمالية", "التأثير", "الخطورة", "خطة التخفيف", "المسؤول"]
for i, h in enumerate(risk_headers, 1):
    ws5.cell(r, i, h)
style_header_row(ws5, r, 7)

risks = [
    (1, "تسريب بيانات الاعتماد (DB/JWT)", "عالي", "حرج", "حرج", "إزالة فورية من الكود، تغيير كلمات المرور، استخدام env vars فقط", "Backend Lead"),
    (2, "عدم تسجيل أرباح المقاولين", "مؤكد", "عالي", "حرج", "إضافة auto-earning في owner-decision endpoint + اختبار", "Backend"),
    (3, "فقدان ملفات المشاريع", "عالي", "عالي", "عالي", "ربط Cloudinary بشكل كامل، نسخ احتياطي تلقائي", "DevOps"),
    (4, "بطء الأداء مع زيادة البيانات", "متوسط", "عالي", "عالي", "إضافة pagination، indexes على الجداول المهمة", "Backend"),
    (5, "عدم عمل التطبيق على الموبايل", "مؤكد", "متوسط", "متوسط", "تصميم responsive في Sprint 3", "Frontend"),
    (6, "صعوبة صيانة App.jsx (5321 سطر)", "عالي", "متوسط", "متوسط", "تقسيم الملف في Sprint 3", "Frontend"),
    (7, "عدم وجود اختبارات", "مؤكد", "متوسط", "متوسط", "إعداد Jest + CI في Sprint 3", "QA"),
    (8, "عدم توافق WCAG", "مؤكد", "منخفض", "منخفض", "مراجعة accessibility في Sprint 3", "Frontend"),
]

for risk in risks:
    r += 1
    for i, val in enumerate(risk, 1):
        ws5.cell(r, i, val)
    style_data_row(ws5, r, 7)
    severity = risk[4]
    if severity == "حرج": ws5.cell(r, 5).fill = redFill
    elif severity == "عالي": ws5.cell(r, 5).fill = PatternFill("solid", fgColor="FFD7D7")
    elif severity == "متوسط": ws5.cell(r, 5).fill = yellowFill
    else: ws5.cell(r, 5).fill = greenFill

ws5.column_dimensions['A'].width = 5
ws5.column_dimensions['B'].width = 32
ws5.column_dimensions['C'].width = 12
ws5.column_dimensions['D'].width = 12
ws5.column_dimensions['E'].width = 12
ws5.column_dimensions['F'].width = 50
ws5.column_dimensions['G'].width = 14

# ══════════════════════════════════════════════════════════
# SHEET 6: API Endpoint Inventory
# ══════════════════════════════════════════════════════════
ws6 = wb.create_sheet("جرد الـ API")
ws6.sheet_view.rightToLeft = True

ws6.merge_cells("A1:G1")
ws6.cell(1,1, "جرد جميع الـ API Endpoints").font = Font(name="Arial", bold=True, size=16, color=WHITE)
ws6.cell(1,1).fill = navyFill; ws6.cell(1,1).alignment = centerA

r = 3
api_headers = ["Method", "Endpoint", "الملف", "الوصف", "Auth", "Role", "الحالة"]
for i, h in enumerate(api_headers, 1):
    ws6.cell(r, i, h)
style_header_row(ws6, r, 7)

endpoints = [
    ("POST", "/api/auth/register", "auth.js", "تسجيل مستخدم جديد", "لا", "عام", "مكتمل"),
    ("POST", "/api/auth/login", "auth.js", "تسجيل الدخول", "لا", "عام", "مكتمل"),
    ("POST", "/api/auth/refresh", "auth.js", "تجديد التوكن", "نعم", "عام", "مكتمل"),
    ("POST", "/api/auth/verify-email", "auth.js", "تفعيل البريد", "لا", "عام", "مكتمل"),
    ("POST", "/api/auth/forgot-password", "auth.js", "نسيت كلمة المرور", "لا", "عام", "مكتمل"),
    ("POST", "/api/auth/reset-password", "auth.js", "إعادة تعيين كلمة المرور", "لا", "عام", "مكتمل"),
    ("PUT", "/api/auth/profile", "auth.js", "تحديث الملف الشخصي", "نعم", "عام", "مكتمل"),
    ("PUT", "/api/auth/change-password", "auth.js", "تغيير كلمة المرور", "نعم", "عام", "مكتمل"),
    ("POST", "/api/auth/logout", "auth.js", "تسجيل الخروج", "نعم", "عام", "مكتمل"),
    ("GET", "/api/projects", "projects.js", "قائمة المشاريع", "نعم", "عام", "مكتمل"),
    ("GET", "/api/projects/:id", "projects.js", "تفاصيل مشروع", "نعم", "عام", "مكتمل"),
    ("POST", "/api/projects", "projects.js", "إنشاء مشروع", "نعم", "owner", "مكتمل"),
    ("POST", "/api/projects/ai-analyze", "projects.js", "تحليل ملف بالـ AI", "نعم", "owner", "مكتمل"),
    ("POST", "/api/projects/ai-create", "projects.js", "إنشاء مشروع من AI", "نعم", "owner", "مكتمل"),
    ("POST", "/api/projects/:id/quotation", "projects.js", "تقديم عرض سعر", "نعم", "contractor", "مكتمل"),
    ("POST", "/api/projects/quotations/:id/accept", "projects.js", "قبول عرض سعر", "نعم", "owner", "مكتمل"),
    ("POST", "/api/projects/:id/apply-inspector", "projects.js", "ترشيح مفتش", "نعم", "inspector", "مكتمل"),
    ("POST", "/api/projects/inspector-apps/:id/accept", "projects.js", "قبول مفتش", "نعم", "owner", "مكتمل"),
    ("POST", "/api/projects/items/:id/submit", "projects.js", "تسليم بند (مقاول)", "نعم", "contractor", "مكتمل"),
    ("POST", "/api/projects/items/:id/review", "projects.js", "فحص بند (مفتش)", "نعم", "inspector", "مكتمل"),
    ("POST", "/api/projects/items/:id/owner-decision", "projects.js", "موافقة/رفض (مالك)", "نعم", "owner", "مكتمل"),
    ("POST", "/api/projects/bank-request", "projects.js", "طلب تمويل بنكي", "نعم", "owner", "مكتمل"),
    ("GET", "/api/projects/bank-requests", "projects.js", "طلباتي البنكية", "نعم", "owner", "مكتمل"),
    ("GET", "/api/projects/completed", "projects.js", "المشاريع المكتملة", "لا", "عام", "مكتمل"),
    ("GET", "/api/dashboard", "dashboard.js", "إحصائيات حسب الدور", "نعم", "عام", "مكتمل"),
    ("GET", "/api/dashboard/contract/:id", "dashboard.js", "تفاصيل العقد", "نعم", "عام", "مكتمل"),
    ("GET", "/api/wallet", "wallet.js", "المحفظة والمعاملات", "نعم", "عام", "مكتمل"),
    ("POST", "/api/wallet/deposit", "wallet.js", "إيداع", "نعم", "عام", "مكتمل"),
    ("GET", "/api/notifications", "notifications.js", "الإشعارات", "نعم", "عام", "مكتمل"),
    ("POST", "/api/notifications/read-all", "notifications.js", "قراءة الكل", "نعم", "عام", "مكتمل"),
    ("GET", "/api/contractors", "contractors.js", "قائمة المقاولين", "نعم", "عام", "مكتمل"),
    ("GET", "/api/inspectors", "inspectors.js", "قائمة المفتشين", "نعم", "عام", "مكتمل"),
    ("GET", "/api/achievements", "achievements.js", "الإنجازات", "نعم", "عام", "مكتمل"),
    ("POST", "/api/achievements/rate", "achievements.js", "تقييم", "نعم", "owner", "مكتمل"),
    ("GET", "/api/items/:id/comments", "comments.js", "تعليقات البند", "نعم", "عام", "مكتمل"),
    ("POST", "/api/items/:id/comments", "comments.js", "إضافة تعليق", "نعم", "عام", "مكتمل"),
    ("POST", "/api/financing/apply", "financing.js", "طلب تمويل", "نعم", "owner", "جزئي"),
    ("GET", "/api/financing/my", "financing.js", "طلباتي", "نعم", "owner", "جزئي"),
    ("GET", "/api/admin/stats", "admin.js", "إحصائيات النظام", "نعم", "admin", "مكتمل"),
    ("GET", "/api/admin/users", "admin.js", "قائمة المستخدمين", "نعم", "admin", "مكتمل"),
    ("PATCH", "/api/admin/users/:id", "admin.js", "تعديل مستخدم", "نعم", "admin", "مكتمل"),
    ("GET", "/api/admin/projects", "admin.js", "قائمة المشاريع", "نعم", "admin", "مكتمل"),
    ("GET", "/api/admin/bank-requests", "admin.js", "طلبات البنك", "نعم", "admin", "مفقود"),
]

for ep in endpoints:
    r += 1
    for i, val in enumerate(ep, 1):
        ws6.cell(r, i, val)
    style_data_row(ws6, r, 7)
    status = ep[6]
    if status == "مكتمل": ws6.cell(r, 7).fill = greenFill
    elif status == "جزئي": ws6.cell(r, 7).fill = yellowFill
    else: ws6.cell(r, 7).fill = redFill

ws6.column_dimensions['A'].width = 8
ws6.column_dimensions['B'].width = 40
ws6.column_dimensions['C'].width = 16
ws6.column_dimensions['D'].width = 28
ws6.column_dimensions['E'].width = 8
ws6.column_dimensions['F'].width = 12
ws6.column_dimensions['G'].width = 10

# Save
output_path = r"C:\Users\Use\Desktop\hassan project 2\FIRST_TOUCH_Gap_Analysis.xlsx"
wb.save(output_path)
print(f"Saved to {output_path}")
